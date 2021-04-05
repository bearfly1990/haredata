import pandas as pd
import openpyxl
import re
import os
from app.common.io import PathUtils
from app.common.log import logger
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from functools import reduce

CROPS_LIST = []
AREA_LIST = ['DSMC', 'DSDM', 'QXDM', 'QXMC', 'CUNDM', 'CUNMC']

AREA_CROPS_LIST = AREA_LIST + CROPS_LIST


class OpenpyxlHelper:
    align_left = Alignment(
        horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(
        horizontal='right', vertical='center', wrap_text=True)

    def __init__(self):
        pass

    @classmethod
    def get_column_letter_from_index(cls, val_int):
        return get_column_letter(val_int)

    @classmethod
    def get_column_index_from_str(cls, val_str):
        return column_index_from_string(val_str)

    @classmethod
    def get_column_letter_from_str_by_diff(cls, val_str, diff):
        return cls.get_column_letter_from_index(cls.get_column_index_from_str(val_str) + diff)

    @classmethod
    def get_column_index_from_str_by_diff(cls, val_str, diff):
        return cls.get_column_index_from_str(val_str) + diff


class ReadHelper:
    def __init__(self):
        pass

    def read_input(self, file):
        global CROPS_LIST
        global AREA_CROPS_LIST

        base_dir = PathUtils.get_dir_name_from_full_path(file)

        # df_sheet_weight = pd.read_excel(
        #     file,
        #     converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)

        df_sheets = {}

        try:
            df_sheets['权重'] = pd.read_excel(
                os.path.join(base_dir, '权重.xlsx'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        except Exception as e:
            df_sheets['权重'] = pd.read_excel(
                os.path.join(base_dir, '权重.xls'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        finally:
            pass

        try:
            df_sheets['裁剪'] = pd.read_excel(
                os.path.join(base_dir, '裁剪.xlsx'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        except Exception as e:
            df_sheets['裁剪'] = pd.read_excel(
                os.path.join(base_dir, '裁剪.xls'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        finally:
            pass

        try:
            df_sheets['未裁剪'] = pd.read_excel(
                os.path.join(base_dir, '未裁剪.xlsx'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        except Exception as e:
            df_sheets['未裁剪'] = pd.read_excel(
                os.path.join(base_dir, '未裁剪.xls'),
                converters={'DSDM': str, 'QXDM': str, 'CUNDM': str}).dropna(axis=0)
        finally:
            pass

        # print(df_sheet_weight)
        # os.path.join(base_dir,'权重.xlsx')
        # df_sheets = []

        # df_sheets = pd.read_excel(
        #     file, sheet_name=['权重', '未裁剪', '裁剪'], engine='openpyxl',
        #     converters={'DSDM': str, 'QXDM': str, 'CUNDM': str})

        df_sheet_weight = df_sheets['权重']

        CROPS_LIST = list(df_sheets['裁剪'].columns)
        for area in AREA_LIST:
            if area in CROPS_LIST:
                CROPS_LIST.remove(area)

        AREA_CROPS_LIST = AREA_LIST + CROPS_LIST

        df_sheet_clip = df_sheets['裁剪'][
            ['CUNDM'] + CROPS_LIST].dropna(axis=0)

        df_sheet_unclipped = df_sheets['未裁剪'][
            ['CUNDM'] + CROPS_LIST].dropna(axis=0)

        dfs = [df_sheet_clip, df_sheet_weight]
        df_sheet_clip = reduce(lambda left, right: pd.merge(
            left, right, on='CUNDM', how='left'), dfs)
        df_sheet_clip = df_sheet_clip[
            ['DSDM', 'QXDM', 'CUNDM'] + CROPS_LIST].dropna(axis=0)

        dfs = [df_sheet_unclipped, df_sheet_weight]
        df_sheet_unclipped = reduce(lambda left, right: pd.merge(
            left, right, on='CUNDM', how='left'), dfs)
        df_sheet_unclipped = df_sheet_unclipped[
            ['DSDM', 'QXDM', 'CUNDM'] + CROPS_LIST].dropna(axis=0)
        # print(df_sheet_unclipped.columns)
        # exit()
        # df_sheet_clip = pd.merge(df_sheet_clip, df_sheet_weight, how='left', on=['CUNDM'], left_on=None,
        #                       right_on=None,
        #                       left_index=False, right_index=False, sort=False,
        #                       suffixes=('_x', '_y'), copy=True, indicator=False)

        return {'权重': df_sheet_weight, '裁剪': df_sheet_clip, '未裁剪': df_sheet_unclipped, 'CROPS': CROPS_LIST}


class DataProcessor:
    style = OpenpyxlHelper()
    reader = ReadHelper()

    def __init__(self):
        self.crops_selected = None
        self.field_input_file = None
        self.field_input_files = None
        self.field_input_dir = None
        self.df_sheet_weight = None
        self.df_sheet_clip = None
        self.df_sheet_unclipped = None

        self.df_process_clip_province = None
        self.df_process_clip_province_summary = None
        self.df_process_unclipped_province = None

        self.df_process_clip_city = None
        self.df_process_unclipped_city = None

        self.df_process_clip_county = None
        self.df_process_unclipped_county = None

        self.df_crops_summary = None
        self.df_crops_summary_city = None

        self.df_crops_summary_province_clipped = None
        self.df_crops_summary_province_unclipped = None

        self.df_crops_summary_city_clipped = None
        self.df_crops_summary_city_unclipped = None

        self.df_crops_summary_county_clipped = None
        self.df_crops_summary_county_unclipped = None

        self.df_process_clip_city_summary = None
        self.df_process_unclipped_city_summary = None

    # ataFrame.to_excel(excel_writer, sheet_name='Sheet1', na_rep='', float_format=None, columns=None, header=True,
    # index=True, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True,
    # encoding=None, inf_rep='inf', verbose=True, freeze_panes=None)
    # def read_input(self):
    #     global CROPS_LIST
    #     global AREA_CROPS_LIST
    #     df_sheets = pd.read_excel(
    #         self.field_input_file, sheet_name=['权重', '未裁剪', '裁剪'], engine='openpyxl',
    #         converters={'DSDM': str, 'QXDM': str, 'CUNDM': str})
    #     # self.input_sheet_name = list(df_sheets)[0]
    #     # if len(list(df_sheets)) > 1:
    #     #     logger.error('文件sheet只能有一个，请选择正确的文件')
    #     #     return
    #
    #     self.df_sheet_weight = df_sheets['权重'].dropna(axis=0)
    #
    #     CROPS_LIST = list(df_sheets['裁剪'].columns)
    #     for area in AREA_LIST:
    #         CROPS_LIST.remove(area)
    #
    #     AREA_CROPS_LIST = AREA_LIST + CROPS_LIST
    #
    #     self.df_sheet_clip = df_sheets['裁剪'][
    #         ['DSDM', 'QXDM', 'CUNDM'] + CROPS_LIST].dropna(axis=0)
    #     self.df_sheet_unclipped = df_sheets['未裁剪'][
    #         ['DSDM', 'QXDM', 'CUNDM'] + CROPS_LIST].dropna(axis=0)
    #     return df_sheets

    def generate_output(self):
        new_file = PathUtils.add_flag_to_file_name(
            self.field_input_file, '计算过程')

        with pd.ExcelWriter(new_file) as writer:
            self.df_sheet_weight.to_excel(writer, sheet_name='权重')
            self.df_process_clip_province.to_excel(writer, sheet_name='省级裁剪')
            self.df_process_unclipped_province.to_excel(
                writer, sheet_name='省级未裁剪')
            self.df_process_clip_city.to_excel(writer, sheet_name='市级裁剪')
            self.df_process_unclipped_city.to_excel(writer, sheet_name='市级未裁剪')
            self.df_crops_summary_city.to_excel(writer, sheet_name='合计（城市）')
            self.df_crops_summary.to_excel(writer, sheet_name='合计（省市）')

        logger.info(f'新文件：{new_file} 已生成！')

        new_file = PathUtils.add_flag_to_file_name(
            self.field_input_file, '处理后')

        with pd.ExcelWriter(new_file) as writer:
            self.df_sheet_weight.to_excel(writer, sheet_name='权重', index=False, columns=['DSMC', 'DSDM', 'QXDM', 'QXMC',
                                                                                         'CUNDM', 'CUNMC', 'W省', 'W市',
                                                                                         'W县'])

            self.df_process_clip_province.to_excel(
                writer, sheet_name='省级裁剪', index=False, columns=AREA_CROPS_LIST)
            self.df_process_unclipped_province.to_excel(writer, sheet_name='省级未裁剪', index=False,
                                                        columns=AREA_CROPS_LIST)
            self.df_process_clip_city.to_excel(
                writer, sheet_name='市级裁剪', index=False, columns=AREA_CROPS_LIST)
            self.df_process_unclipped_city.to_excel(
                writer, sheet_name='市级未裁剪', index=False, columns=AREA_CROPS_LIST)
            self.df_process_clip_county.to_excel(
                writer, sheet_name='县级裁剪', index=False, columns=AREA_CROPS_LIST)
            self.df_process_unclipped_county.to_excel(
                writer, sheet_name='县级未裁剪', index=False, columns=AREA_CROPS_LIST)
            # print(self.df_crops_summary_city.columns)
            # print(self.df_crops_summary)

            self.df_crops_summary_city.to_excel(writer, sheet_name='合计（城市）', index=False,
                                                columns=['level', 'clip_flag', 'DSDM', 'DSMC'] + CROPS_LIST)

            # print(self.df_crops_summary.columns)
            # print(['level', 'clip_flag'] + CROPS_LIST)
            self.df_crops_summary.to_excel(writer, sheet_name='合计（省市）', index=False,
                                           columns=['level', 'clip_flag'] + CROPS_LIST)

        logger.info(f'新文件：{new_file} 已生成！')

    # @staticmethod
    # def write_header(sheet_summary, value_list, row_start, col_start_char) -> int:
    #     row_current = row_start
    #     for i, value in enumerate(value_list):
    #         cell = sheet_summary.cell(row=row_current,
    #                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + i)
    #         cell.value = value
    #     return OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, len(value_list))

    @staticmethod
    def write_summary(sheet_summary, df_summary, row_start, col_start_char, diff):
        col_start_char = OpenpyxlHelper.get_column_letter_from_str_by_diff(
            col_start_char, diff)
        row_current = row_start
        for index, row in df_summary.iterrows():
            for i, crop in enumerate(CROPS_LIST):
                cell = sheet_summary.cell(row=row_current,
                                          column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + i)
                cell.value = row[i]
                cell.number_format = '#,##0.00'

            cell = sheet_summary.cell(
                row=row_current, column=OpenpyxlHelper.get_column_index_from_str_by_diff(
                    col_start_char, -1)
            )
            start_cell = f'{col_start_char}{row_current}'
            ended_cell = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, len(CROPS_LIST) - 1)}{row_current}'
            cell.value = f'=sum({start_cell}:{ended_cell})'
            cell.number_format = '#,##0.00'
            row_current = row_current + 1

    def write_summary_city_sum_crops_factor(self, sheet_summary, df_summary, row_start, col_start_char, diff):
        # col_start_char = OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, diff)
        row_current = row_start
        for index, row in df_summary.iterrows():
            cells_crops_selected_clipped = []
            cells_crops_selected_unclipped = []

            for i, crop in enumerate(CROPS_LIST):
                cell = sheet_summary.cell(
                    row=row_current,
                    column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + i)

                if df_summary.columns[i + 2] in self.crops_selected:
                    cell_str = f'{cell.column_letter}{cell.row}'
                    cells_crops_selected_clipped.append(cell_str)
                    # cells_crops_selected_clipped.append(sheet_summary[cell_str].value)
                    cell_str = f'{cell.column_letter}{cell.row + diff}'
                    cells_crops_selected_unclipped.append(cell_str)
                    # cells_crops_selected_unclipped.append(sheet_summary[cell_str].value)
            cell = sheet_summary.cell(
                row=row_current,
                column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + len(CROPS_LIST))

            cell.value = f"=sum({'+'.join(cells_crops_selected_unclipped)})/sum({'+'.join(cells_crops_selected_clipped)})"
            # cell.number_format = '#,##0.0000'
            sheet_summary.column_dimensions[
                OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3 + len(CROPS_LIST))].width = 20
            # cell.value = sum(cells_crops_selected_clipped) / sum(cells_crops_selected_unclipped)
            row_current = row_current + 1

    def write_summary_city_sum(self, sheet_summary, df_summary, row_start, col_start_char, diff):
        col_start_char = OpenpyxlHelper.get_column_letter_from_str_by_diff(
            col_start_char, diff)
        row_current = row_start
        for index, row in df_summary.iterrows():
            sheet_summary.cell(row=row_current,
                               column=OpenpyxlHelper.get_column_index_from_str(col_start_char)).value = row[0]

            sheet_summary.cell(row=row_current,
                               column=OpenpyxlHelper.get_column_index_from_str_by_diff(
                                   col_start_char, 1)).value = row[1]

            start_cell = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3)}{row_current}'
            ended_cell = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3 + len(CROPS_LIST))}{row_current}'
            formula = f'=sum({start_cell}:{ended_cell})'
            cell = sheet_summary.cell(row=row_current,
                                      column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 2)
            cell.value = formula
            cell.number_format = '#,##0.00'

            for i, crop in enumerate(CROPS_LIST):
                cell = sheet_summary.cell(
                    row=row_current,
                    column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + i)
                cell.value = row[i + 2]
                cell.number_format = '#,##0.00'
                # if df_summary.columns[i+2] in self.crops_selected:
                #     cell_str = f'{cell.column_letter}{cell.row}'
                # print(cell_str, sheet_summary[cell_str].value)
                # sum_crops_selected = sum_crops_selected + row[i + 2]
            # print(sum_crops_selected)
            row_current = row_current + 1

    def write_summary_county_sum(self, sheet_summary, df_summary, row_start, col_start_char, diff):
        col_start_char = OpenpyxlHelper.get_column_letter_from_str_by_diff(
            col_start_char, diff)
        row_current = row_start
        city_row = row_start
        city_code_current = None
        # print(df_summary)
        for index, row in df_summary.iterrows():
            # print(index, )
            if index[0] != city_code_current:
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char)).value = index[0]
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(
                                       col_start_char)).alignment = self.style.align_left

                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 1).value = index[2]

                city_code_current = index[0]

                for i, crop in enumerate(['total'] + CROPS_LIST):
                    cell_start = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, i + 2)}{city_row + 1}'
                    cell_ended = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, i + 2)}{row_current - 1}'
                    cell = sheet_summary.cell(row=city_row,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 2)
                    cell.value = f'=sum({cell_start}:{cell_ended})'
                    cell.number_format = '#,##0.00'
                city_row = row_current
                row_current = row_current + 1
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char)).value = index[1]
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(
                                       col_start_char)).alignment = self.style.align_right

                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 1).value = index[3]

                cell_start = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3)}{row_current}'
                cell_ended = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, len(CROPS_LIST) - 1 + 3)}{row_current}'

                cell = sheet_summary.cell(row=row_current, column=OpenpyxlHelper.get_column_index_from_str(
                    col_start_char) + 2)
                cell.value = f'=sum({cell_start}:{cell_ended})'
                cell.number_format = '#,##0.00'

                for i, crop in enumerate(CROPS_LIST):
                    cell = sheet_summary.cell(row=row_current,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 3)
                    cell.value = row[i]
                    cell.number_format = '#,##0.00'
            else:
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char)).value = index[1]
                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(
                                       col_start_char)).alignment = self.style.align_right

                sheet_summary.cell(row=row_current,
                                   column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 1).value = index[3]

                cell_start = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3)}{row_current}'
                cell_ended = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, len(CROPS_LIST) - 1 + 3)}{row_current}'

                cell = sheet_summary.cell(row=row_current, column=OpenpyxlHelper.get_column_index_from_str(
                    col_start_char) + 2)
                cell.value = f'=sum({cell_start}:{cell_ended})'
                cell.number_format = '#,##0.00'

                for i, crop in enumerate(CROPS_LIST):
                    cell = sheet_summary.cell(row=row_current,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 3)
                    cell.value = row[i]
                    cell.number_format = '#,##0.00'
            row_current = row_current + 1

        for i, crop in enumerate(['total'] + CROPS_LIST):
            cell_start = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, i + 2)}{city_row + 1}'
            cell_ended = f'{OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, i + 2)}{row_current - 1}'
            cell = sheet_summary.cell(row=city_row,
                                      column=OpenpyxlHelper.get_column_index_from_str(
                                          col_start_char) + i + 2)
            cell.value = f'=sum({cell_start}:{cell_ended})'
            cell.number_format = '#,##0.00'

    def write_summary_county_sum_crops_factor(self, sheet_summary, df_summary, row_start, col_start_char, diff):
        # col_start_char = OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, diff)
        row_current = row_start
        city_row = row_start
        city_code_current = None
        # print(df_summary)
        for index, row in df_summary.iterrows():
            # print(index, )
            cells_crops_selected_clipped = []
            cells_crops_selected_unclipped = []
            if index[0] != city_code_current:
                # 第一个市
                city_row = row_current
                city_code_current = index[0]
                for i, crop in enumerate(CROPS_LIST):
                    cell = sheet_summary.cell(row=city_row,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 3)
                    if df_summary.columns[i] in self.crops_selected:
                        cell_str = f'{cell.column_letter}{cell.row}'
                        cells_crops_selected_clipped.append(cell_str)
                        cell_str = f'{cell.column_letter}{cell.row + diff}'
                        cells_crops_selected_unclipped.append(cell_str)

                cell = sheet_summary.cell(
                    row=city_row,
                    column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + len(CROPS_LIST))

                cell.value = f"=sum({'+'.join(cells_crops_selected_unclipped)})/sum({'+'.join(cells_crops_selected_clipped)})"
                # cell.number_format = '#,##0.0000'
                sheet_summary.column_dimensions[
                    OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3 + len(CROPS_LIST))].width = 20

                row_current = row_current + 1
                # 第一个县区
                cells_crops_selected_clipped = []
                cells_crops_selected_unclipped = []
                for i, crop in enumerate(CROPS_LIST):
                    cell = sheet_summary.cell(row=row_current,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 3)
                    if df_summary.columns[i] in self.crops_selected:
                        cell_str = f'{cell.column_letter}{cell.row}'
                        cells_crops_selected_clipped.append(cell_str)
                        cell_str = f'{cell.column_letter}{cell.row + diff}'
                        cells_crops_selected_unclipped.append(cell_str)

                cell = sheet_summary.cell(
                    row=row_current,
                    column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + len(CROPS_LIST))

                cell.value = f"=sum({'+'.join(cells_crops_selected_unclipped)})/sum({'+'.join(cells_crops_selected_clipped)})"
                # cell.number_format = '#,##0.0000'
                sheet_summary.column_dimensions[
                    OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3 + len(CROPS_LIST))].width = 20

                row_current = row_current + 1
            else:
                cells_crops_selected_clipped = []
                cells_crops_selected_unclipped = []
                for i, crop in enumerate(CROPS_LIST):
                    cell = sheet_summary.cell(row=row_current,
                                              column=OpenpyxlHelper.get_column_index_from_str(
                                                  col_start_char) + i + 3)

                    if df_summary.columns[i] in self.crops_selected:
                        cell_str = f'{cell.column_letter}{cell.row}'
                        cells_crops_selected_clipped.append(cell_str)
                        cell_str = f'{cell.column_letter}{cell.row + diff}'
                        cells_crops_selected_unclipped.append(cell_str)
                cell = sheet_summary.cell(
                    row=row_current,
                    column=OpenpyxlHelper.get_column_index_from_str(col_start_char) + 3 + len(CROPS_LIST))

                cell.value = f"=sum({'+'.join(cells_crops_selected_unclipped)})/sum({'+'.join(cells_crops_selected_clipped)})"
                # cell.number_format = '#,##0.0000'
                sheet_summary.column_dimensions[
                    OpenpyxlHelper.get_column_letter_from_str_by_diff(col_start_char, 3 + len(CROPS_LIST))].width = 20

                row_current = row_current + 1

    @classmethod
    def write_header_label(cls, sheet_summary, str_list, col_start_str, row_start, axis=0):
        col_start_index = OpenpyxlHelper.get_column_index_from_str(
            col_start_str)
        if axis == 0:
            for i, value in enumerate(str_list):
                sheet_summary.cell(
                    row=row_start, column=col_start_index).value = value
                col_start_index = col_start_index + 1
        elif axis == 1:
            for i, value in enumerate(str_list):
                sheet_summary.cell(
                    row=row_start, column=col_start_index).value = value
                row_start = row_start + 1

    @classmethod
    def write_province_header_label(cls, sheet_summary, diff):
        cls.write_header_label(
            sheet_summary, ['裁剪', '省权重', '市权重合计', '县权重合计'], 'A', 1, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST, 'D', 1, axis=0)
        cls.write_header_label(sheet_summary, ['小品种调整因子'],
                               OpenpyxlHelper.get_column_letter_from_str_by_diff('D', len(CROPS_LIST)), 1, axis=0)
        cls.write_header_label(
            sheet_summary, ['未裁剪', '省权重', '市权重合计', '县权重合计'], 'A', 1 + diff, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1 + diff, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST,
                               'D', 1 + diff, axis=0)

        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 1)
        # cls.write_header_label(sheet_summary, ['裁剪', '省权重', '市权重合计', '县权重合计'], col_str, 1, axis=1)
        #
        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 3)
        # cls.write_header_label(sheet_summary, ['合计'], col_str, 1, axis=1)
        #
        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 4)
        # cls.write_header_label(sheet_summary, CROPS_LIST, col_str, 1, axis=0)

    @classmethod
    def write_province_header_label_fixed(cls, sheet_summary, diff):
        cls.write_header_label(
            sheet_summary, ['市上报数据', '市权重合计', '县权重合计'], 'A', 1, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST, 'D', 1, axis=0)
        # cls.write_header_label(sheet_summary, ['小品种调整因子'],
        #                        OpenpyxlHelper.get_column_letter_from_str_by_diff('D', len(CROPS_LIST)), 1, axis=0)
        cls.write_header_label(
            sheet_summary, ['调整因子', '市权重合计', '县权重合计'], 'A', 1 + diff + 2, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1 + diff + 2, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST,
                               'D', 1 + diff + 2, axis=0)

        cls.write_header_label(
            sheet_summary, ['裁剪', '市权重合计', '县权重合计'], 'A', 1 + 2 * diff + 4, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1 + 2 * diff + 4, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST,
                               'D', 1 + 2 * diff + 4, axis=0)

    @classmethod
    def write_city_header_label(cls, sheet_summary, row_diff):
        cls.write_header_label(
            sheet_summary, ['裁剪', '市权重合计', '县权重合计'], 'A', 1, axis=1)
        cls.write_header_label(sheet_summary, ['合计'], 'C', 1, axis=1)
        cls.write_header_label(sheet_summary, CROPS_LIST, 'D', 1, axis=0)
        cls.write_header_label(sheet_summary, ['小品种调整因子'],
                               OpenpyxlHelper.get_column_letter_from_str_by_diff('D', len(CROPS_LIST)), 1, axis=0)
        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 1)
        cls.write_header_label(
            sheet_summary, ['未裁剪', '市权重合计', '县权重合计'], 'A', 1 + row_diff, axis=1)
        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 3)
        cls.write_header_label(
            sheet_summary, ['合计'], 'C', 1 + row_diff, axis=1)
        # col_str = OpenpyxlHelper.get_column_letter_from_str_by_diff('C', len(CROPS_LIST) + 4)
        cls.write_header_label(sheet_summary, CROPS_LIST,
                               'D', 1 + row_diff, axis=0)

    @classmethod
    def auto_width(cls, sheet):
        # width = 2.0
        # height = width * (2.2862 / 0.3612)
        # print("row:", sheet_summary_city.max_row, "column:", sheet_summary_city.max_column)

        for i in range(1, sheet.max_column + 1):
            max_width = 13
            for j in range(1, sheet.max_row + 1):
                cell = f'{OpenpyxlHelper.get_column_letter_from_index(i)}{j}'
                # height = sheet_summary_city.row_dimensions[j].height
                # width = sheet.column_dimensions[OpenpyxlHelper.get_column_letter_from_index(i)].width
                value_width = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(sheet[cell].value))) + len(
                    str(sheet[cell].value))
                # value_width = len(str(sheet_summary_city[cell].value))
                if value_width > max_width:
                    max_width = value_width
                # print(cell, width, value_width, max_width)
            sheet.column_dimensions[OpenpyxlHelper.get_column_letter_from_index(
                i)].width = max_width + 2

    def write_df_to_sheet(self, sheet, df_sheet):
        row_current = 1
        for i, column in enumerate(df_sheet.columns):
            sheet.cell(row=row_current, column=i + 1).value = column
        row_current = row_current + 1
        for i, row in df_sheet.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=i + row_current, column=j + 1).value = value

    def generate_report(self):
        new_file = PathUtils.add_flag_to_file_name(self.field_input_file, '报表')
        # workbook = openpyxl.load_workbook('./template/模板.xlsx')
        workbook = openpyxl.Workbook()
        sheet_weight = workbook.create_sheet('权重输出')
        sheet_summary = workbook.create_sheet('省级权限输出界面')
        sheet_summary_fixed = workbook.create_sheet('调整后省级输出')
        sheet_summary_city = workbook.create_sheet('市级权限输出界面')
        sheet_summary_city_fixed = workbook.create_sheet('调整后市级输出')
        workbook.remove(workbook['Sheet'])
        self.df_sheet_weight = self.df_sheet_weight[
            ['DSMC', 'DSDM', 'QXDM', 'QXMC', 'CUNDM', 'CUNMC', 'W省', 'W市', 'W县']]
        self.write_df_to_sheet(sheet_weight, self.df_sheet_weight)
        row_diff_province = len(self.df_process_unclipped_city_summary) + 6
        self.write_province_header_label(sheet_summary, row_diff_province)

        self.write_summary(
            sheet_summary, self.df_crops_summary_province_clipped, 2, 'D', 0)
        self.write_summary(
            sheet_summary, self.df_crops_summary_province_unclipped, 2 + row_diff_province, 'D', 0)

        self.write_summary(
            sheet_summary, self.df_crops_summary_city_clipped, 3, 'D', 0)
        self.write_summary(
            sheet_summary, self.df_crops_summary_city_unclipped, 3 + row_diff_province, 'D', 0)
        self.write_summary(
            sheet_summary, self.df_crops_summary_county_clipped, 4, 'D', 0)
        self.write_summary(
            sheet_summary, self.df_crops_summary_county_unclipped, 4 + row_diff_province, 'D', 0)

        self.write_summary_city_sum(
            sheet_summary, self.df_process_clip_city_summary, 5, 'A', 0)
        self.write_summary_city_sum(sheet_summary, self.df_process_unclipped_city_summary, 5 + row_diff_province, 'A',
                                    0)
        # sheet_summary_fixed
        row_diff_province = len(self.df_crops_summary_city_clipped)
        self.write_province_header_label_fixed(sheet_summary_fixed, row_diff_province + 3)
        # self.write_summary(
        #     sheet_summary_fixed, self.df_crops_summary_province_clipped, 2 + 2*(row_diff_province + 2), 'D', 0)





        self.write_summary(
            sheet_summary_fixed, self.df_crops_summary_city_clipped, 2 + 2 * row_diff_province + 4, 'D', 0)
        self.write_summary(
            sheet_summary_fixed, self.df_crops_summary_county_clipped, 3 + 2 * row_diff_province + 4, 'D', 0)
        self.write_summary_city_sum(
            sheet_summary_fixed, self.df_process_clip_city_summary, 4 + 2 * row_diff_province + 4, 'A', 0)




        df_process_unclipped_county_group_by = self.df_process_unclipped_county.groupby(
            by=['DSDM', 'QXDM', 'DSMC', 'QXMC'])[CROPS_LIST].sum()
        df_process_clipped_county_group_by = self.df_process_clip_county.groupby(
            by=['DSDM', 'QXDM', 'DSMC', 'QXMC'])[CROPS_LIST].sum()

        df_process_unclipped_city_group_by = self.df_process_unclipped_county.groupby(
            by=['DSDM', 'DSMC'])[CROPS_LIST].sum()

        row_diff_city = len(df_process_unclipped_county_group_by) + \
                        len(df_process_unclipped_city_group_by) + 5

        self.write_city_header_label(sheet_summary_city, row_diff_city)
        self.write_summary(sheet_summary_city,
                           self.df_crops_summary_city_clipped, 2, 'D', 0)
        self.write_summary(
            sheet_summary_city, self.df_crops_summary_city_unclipped, 2 + row_diff_city, 'D', 0)
        self.write_summary(sheet_summary_city,
                           self.df_crops_summary_county_clipped, 3, 'D', 0)
        self.write_summary(
            sheet_summary_city, self.df_crops_summary_county_unclipped, 3 + row_diff_city, 'D', 0)

        self.write_summary_county_sum(
            sheet_summary_city, df_process_clipped_county_group_by, 4, 'A', 0)
        self.write_summary_county_sum(sheet_summary_city, df_process_unclipped_county_group_by, 4 + row_diff_city, 'A',
                                      0)
        self.auto_width(sheet_weight)
        self.auto_width(sheet_summary)
        self.auto_width(sheet_summary_city)

        self.write_summary_city_sum_crops_factor(sheet_summary, self.df_process_clip_city_summary, 5, 'A',
                                                 row_diff_province)

        self.write_summary_county_sum_crops_factor(sheet_summary_city, df_process_clipped_county_group_by, 4, 'A',
                                                   row_diff_city)

        workbook.save(filename=new_file)
        logger.info(f'新文件：{new_file} 已生成！')
        # print(self.df_process_unclipped_city)

        # print(type(new_series))
        # print(new_series.columns)
        # df_output['QXDM'] = df_output.index
        # df_output = df_output.reset_index(drop=True)
        # print(df_output)
        # e = df_output.loc[[(3301, 330109, '杭州市', '萧山区')]]  # DataFrame索引取值

    def process_weight(self):
        logger.info('计算C1/WA3...')
        self.df_sheet_weight['C1/WA3'] = self.df_sheet_weight.apply(
            lambda x: self.c1_wa3(x), axis=1)

        logger.info('计算C1/WA2...')
        self.df_sheet_weight['C1/WA2'] = self.df_sheet_weight.apply(
            lambda x: self.c1_wa2(x), axis=1)

        logger.info('计算C1/WA...')
        self.df_sheet_weight['C1/WA'] = self.df_sheet_weight.apply(
            lambda x: self.c1_wa(x), axis=1)

        logger.info('计算n...')
        self.df_sheet_weight['n'] = self.df_sheet_weight.apply(
            lambda x: self.n(x), axis=1)

        logger.info('计算n2...')
        self.df_sheet_weight['n2'] = self.df_sheet_weight.apply(
            lambda x: self.n2(x), axis=1)

        logger.info('计算n3...')
        self.df_sheet_weight['n3'] = self.df_sheet_weight['DSDM'].count()

        logger.info('计算C1county...')
        self.df_sheet_weight['C1county'] = self.df_sheet_weight.apply(
            lambda x: self.c1_county(x), axis=1)

        logger.info('计算C1city...')
        self.df_sheet_weight['C1city'] = self.df_sheet_weight.apply(
            lambda x: self.c1_city(x), axis=1)

        logger.info('计算C1pro...')
        self.df_sheet_weight['C1pro'] = self.df_sheet_weight.apply(
            lambda x: self.c1_pro(x), axis=1)

        logger.info('计算W省..')
        self.df_sheet_weight['W省'] = self.df_sheet_weight.apply(
            lambda x: self.w_province(x), axis=1)

        logger.info('计算W市..')
        self.df_sheet_weight['W市'] = self.df_sheet_weight.apply(
            lambda x: self.w_city(x), axis=1)

        logger.info('计算W县..')
        self.df_sheet_weight['W县'] = self.df_sheet_weight.apply(
            lambda x: self.w_country(x), axis=1)

        # self.df_sheet_weight['W省'] = self.df_sheet_weight['W省'].map(lambda x: ("%.5f") % x)
        # self.df_sheet_weight['W市'] = self.df_sheet_weight['W市'].map(lambda x: ("%.5f") % x)
        # self.df_sheet_weight['W县'] = self.df_sheet_weight['W县'].map(lambda x: ("%.5f") % x)
        # self.df_sheet_weight[['W省', 'W市', 'W县']] = self.df_sheet_weight[['W省', 'W市', 'W县']].astype(float)
        self.df_sheet_weight[['CUNDM']
        ] = self.df_sheet_weight[['CUNDM']].astype(str)

    def process_clip(self, level):
        return self.process_crop(self.df_sheet_clip, level, '裁剪')

    def process_unclipped(self, level):
        return self.process_crop(self.df_sheet_unclipped, level, '未裁剪')

    def process_crop(self, df_sheet_crop, level, clip_flag):
        df_process = pd.merge(self.df_sheet_weight, df_sheet_crop, how='inner', on=['DSDM', 'QXDM',
                                                                                    'CUNDM'], left_on=None,
                              right_on=None,
                              left_index=False, right_index=False, sort=False,
                              suffixes=('_x', '_y'), copy=True, indicator=False)
        # print(self.df_sheet_weight.columns)
        # print(df_sheet_crop.columns)
        # print(df_process.columns)
        for crop in CROPS_LIST:
            logger.info(f'计算{level}级{clip_flag}:{crop}...')
            df_process[crop] = df_process.apply(
                lambda x: x[crop] * x[f'W{level}'], axis=1)

        return df_process

    def process(self):
        df_sheets = self.reader.read_input(self.field_input_file)
        self.df_sheet_weight = df_sheets['权重']
        self.df_sheet_clip = df_sheets['裁剪']
        self.df_sheet_unclipped = df_sheets['未裁剪']

        self.process_weight()

        self.df_process_clip_province = self.process_clip('省')
        self.df_process_unclipped_province = self.process_unclipped('省')

        self.df_process_clip_city = self.process_clip('市')
        self.df_process_unclipped_city = self.process_unclipped('市')

        self.df_process_clip_county = self.process_clip('县')
        self.df_process_unclipped_county = self.process_unclipped('县')

        self.df_crops_summary_province_clipped = self.summary_corp(
            self.df_process_clip_province, '省', '剪裁')
        self.df_crops_summary_province_unclipped = self.summary_corp(
            self.df_process_unclipped_province, '省', '未剪裁')
        self.df_crops_summary_city_clipped = self.summary_corp(
            self.df_process_clip_city, '市', '剪裁')
        self.df_crops_summary_city_unclipped = self.summary_corp(
            self.df_process_unclipped_city, '市', '未剪裁')
        self.df_crops_summary_county_clipped = self.summary_corp(
            self.df_process_clip_county, '县', '剪裁')
        self.df_crops_summary_county_unclipped = self.summary_corp(
            self.df_process_unclipped_county, '县', '未剪裁')

        self.df_crops_summary = pd.concat([self.df_crops_summary_province_clipped,
                                           self.df_crops_summary_province_unclipped,
                                           self.df_crops_summary_city_clipped,
                                           self.df_crops_summary_city_unclipped,
                                           self.df_crops_summary_county_clipped,
                                           self.df_crops_summary_county_unclipped])

        self.df_crops_summary_city = self.summary_crops_city(
            self.df_process_clip_province, '省', '剪裁')
        # print(self.df_crops_summary_city)
        self.df_crops_summary_city = pd.concat([self.df_crops_summary_city,
                                                self.summary_crops_city(self.df_process_unclipped_province, '省',
                                                                        '未剪裁')])
        self.df_process_clip_city_summary = self.summary_crops_city(
            self.df_process_clip_city, '市', '剪裁')
        self.df_crops_summary_city = pd.concat([self.df_crops_summary_city,
                                                self.df_process_clip_city_summary])
        self.df_process_unclipped_city_summary = self.summary_crops_city(
            self.df_process_unclipped_city, '市', '未剪裁')
        self.df_crops_summary_city = pd.concat([self.df_crops_summary_city,
                                                self.df_process_unclipped_city_summary])
        # self.df_crops_summary_city =
        # print(self.df_crops_summary)
        # print(self.df_crops_summary_city)
        # self.generate_output()
        self.generate_report()

    def summary_corp(self, df_process, level, clip_flag):
        data = {}
        for crop in CROPS_LIST:
            data[crop] = [df_process[crop].sum()]

        data['level'] = [level]
        data['clip_flag'] = [clip_flag]
        return pd.DataFrame(data)

    def summary_crops_city(self, df_process, level, clip_flag):
        df_output = None
        for crop in CROPS_LIST:
            if df_output is not None:
                df_output = pd.merge(df_output, df_process.groupby(['DSDM', 'DSMC'])[crop].sum(), how='inner',
                                     on=['DSDM', 'DSMC'])
            else:
                df_output = df_process.groupby(['DSDM', 'DSMC'])[crop].sum()
                df_output = pd.DataFrame(df_output)
                df_output.reset_index(inplace=True)

        df_output['level'] = level
        df_output['clip_flag'] = clip_flag
        # print(df_output)
        return df_output

    def c1_wa3(self, x):
        return x['C1'] / x['C2'] * x['C2pro']

    def c1_wa2(self, x):
        return x['C1'] / x['C2'] * x['C2city']

    def c1_wa(self, x):
        return x['C1'] / x['C2'] * x['C2county']

    def n(self, x):
        # print(x['QXDM'])
        # print(self.df_sheet.loc[:,'QXDM'].value_counts())
        # print(type(self.df_sheet.loc[:,'QXDM'].value_counts()))
        # print(list(self.df_sheet['DSDM']))
        return self.df_sheet_weight.groupby(['QXDM']).size()[x['QXDM']]

    def n2(self, x):
        return self.df_sheet_weight.groupby(['DSDM']).size()[x['DSDM']]

    def c1_county(self, x):
        return self.df_sheet_weight.groupby(['QXDM'])['C1/WA'].sum()[x['QXDM']] / x['n']

    def c1_city(self, x):
        return self.df_sheet_weight.groupby(['DSDM'])['C1/WA2'].sum()[x['DSDM']] / x['n2']

    def c1_pro(self, x):  # C1pro
        return self.df_sheet_weight['C1/WA3'].sum() / x['n3']

    def w_province(self, x):
        if int(x['C1R']) == 0 or int(x['n3']) == 0:
            return 0
        return x['C1pro'] / x['C1R'] / x['n3']

    def w_city(self, x):
        if int(x['C1R']) == 0 or int(x['n3']) == 0:
            return 0
        return x['C1city'] / x['C1R'] / x['n2']

    def w_country(self, x):
        if int(x['C1R']) == 0 or int(x['n3']) == 0:
            return 0
        return x['C1county'] / x['C1R'] / x['n']
